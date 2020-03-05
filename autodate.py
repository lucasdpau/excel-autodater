import openpyxl, datetime

wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
cell_a1 = ws["A1"]
cell_a2 = ws["A2"]
cell_a3 = ws["A3"]
cell_a4 = ws["A4"]
cell_b1 = ws["B1"]

cell_range_col_a = ws["a1:a4"]

def see_if_date_has_passed(date):
    # date needs to be a datetime.datetime object
    print(date)
    return date <= datetime.datetime.today()

def update_date(cell, timedelta):
    #increase date value by (int) timedelta number of days if the day has passed 
    date_increase = datetime.timedelta(days=int(timedelta))
    if see_if_date_has_passed(cell.value):
        cell.value += date_increase
    return cell.value

def check_cell_range(worksheet, cell1, cell2, timedelta):
    #collects all cells in the range between cell1, and cell2
    search_str = f"{cell1}:{cell2}"
    cell_range = worksheet[search_str]
    check_list = []
    for items in cell_range:
        check_list.append(items[0])
    for cells in check_list:
        update_date(cells, timedelta)
    
if __name__ == "main":
    pass

#go through each cell in the range
#format each cell into a workable datetime format
#compare the cell's value to today's datetime format
#if cell > today, leave it. else add 28,30,60, or 90 days
